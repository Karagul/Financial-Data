import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook
from time import sleep

a= ' MMM ABT ABBV ACN ATVI AYI ADBE AAP AES AET AMG AFL A APD AKAM ALK ALB ALXN ALLE AGN ADS LNT ALL GOOGL GOOG MO AMZN AEE AAL AEP AXP AIG AMT AWK AMP ABC AME AMGN APH APC ADI ANTM AON APA AIV AAPL AMAT ADM ARNC AJG AIZ T ADSK ADP AN AZO AVB AVY BHI BLL BAC BCR BAX BBT BDX BBBY BRK.B BBY BIIB BLK HRB BA BWA BXP BSX BMY AVGO BF.B CHRW CA COG CPB COF CAH KMX CCL CAT CBOE CBG CBS CELG CNC CNP CTL CERN CF SCHW CHTR CHK CVX CMG CB CHD CI XEC CINF CTAS CSCO C CFG CTXS CME CMS COH KO CTSH CL CMCSA CMA CAG CXO COP ED STZ GLW COST COTY CCI CSRA CSX CMI CVS DHI DHR DRI DVA DE DLPH DAL XRAY DVN DLR DFS DISCA DISCK DG DLTR D DOV DOW DPS DTE DD DUK DNB ETFC EMN ETN EBAY ECL EIX EW EA EMR ETR EVHC EOG EQT EFX EQIX EQR ESS EL ES EXC EXPE EXPD ESRX EXR XOM FFIV FB FAST FRT FDX FIS FITB FSLR FE FISV FLIR FLS FLR FMC FTI FL F FTV FBHS BEN FCX FTR GPS GRMN GD GE GGP GIS GM GPC GILD GPN GS GT GWW HAL HBI HOG HAR HRS HIG HAS HCA HCP HP HSIC HES HPE HOLX HD HON HRL HST HPQ HUM HBAN IDXX ITW ILMN INCY IR INTC ICE IBM IP IPG IFF INTU ISRG IVZ IRM JBHT JEC SJM JNJ JCI JPM JNPR KSU K KEY KMB KIM KMI KLAC KSS KHC KR LB LLL LH LRCX LEG LEN LUK LVLT LLY LNC LLTC LKQ LMT L LOW LYB MTB MAC M MNK MRO MPC MAR MMC MLM MAS MA MAT MKC MCD MCK MJN MDT MRK MET MTD KORS MCHP MU MSFT MAA MHK TAP MDLZ MON MNST MCO MS MSI MUR MYL NDAQ NOV NAVI NTAP NFLX NWL NFX NEM NWSA NWS NEE NLSN NKE NI NBL JWN NSC NTRS NOC NRG NUE NVDA ORLY OXY OMC OKE ORCL PCAR PH PDCO PAYX PYPL PNR PBCT PEP PKI PRGO PFE PCG PM PSX PNW PXD PNC RL PPG PPL PX PCLN PFG PG PGR PLD PRU PEG PSA PHM PVH QRVO QCOM PWR DGX RRC RTN O RHT REG REGN RF RSG RAI RHI ROK COL ROP ROST RCL R SPGI CRM SCG SLB SNI STX SEE SRE SHW SIG SPG SWKS SLG SNA SO LUV SWN SWK SPLS SBUX STT SRCL SYK STI SYMC SYF SYY TROW TGT TEL TGNA TDC TSO TXN TXT BK CLX COO HSY MOS TRV DIS TMO TIF TWX TJX TMK TSS TSCO TDG RIG TRIP FOXA FOX TSN USB UDR ULTA UA UAA UNP UAL UNH UPS URI UTX UHS UNM URBN VFC VLO VAR VTR VRSN VRSK VZ VRTX VIAB V VNO VMC WMT WBA WM WAT WEC WFC HCN WDC WU WRK WY WHR WFM WMB WLTW WYN WYNN XEL XRX XLNX XL XYL YHOO YUM ZBH ZION ZTS'
# enter tickers here


a=a.split() # makes list 

b=a
url='http://www.msn.com/en-us/money/stockdetails/analysis/fi-126.1.'

wb = Workbook()
sheet=wb.active
sheet.title = 'a'

#define list to append later
debtl=[]
currentl=[]
pbl=[]
perl=[]
bshl=[]
indl =[]
n=0


for a in a:
    n+=1
    print n
    fullurl = url + a + '.NYS' #append full url from list

    
    page =requests.get(fullurl)
    if n==60:
        print 'sleeping'
        sleep(10)
        
    
    if n==200:
        print 'sleeping'
        sleep(10)
        
    if n==300:
        print 'sleeping'
        sleep(10)
        
    if n==400:
        print 'sleeping'
        sleep(10)
        
    
    try:
        #getting data
        soup = BeautifulSoup(page.content, 'lxml')
        debt = soup.find(attrs={'title': 'Debt/Equity Ratio'})
        dval= debt.next_element.next_element.next_element.next_element.next_element.next_element.next_element.next_element
        
        

        dval = dval.text
        
        
    

        currentr = soup.find(attrs={'title': 'Current Ratio'})
        currentrval= currentr.next_element.next_element.next_element.next_element.next_element.next_element.next_element.next_element
        currentval = currentrval.text

        pbv = soup.find(attrs={'title': 'Price/Book Value'})   
        pbval = pbv.next_element.next_element.next_element.next_element
        pbval = pbval.text

        per = soup.find(attrs={'title': 'Current P/E Ratio'})
        perval = per.next_element.next_element.next_element.next_element.next_element.next_element.next_element.next_element
        perval = perval.text

        bsh = soup.find(attrs={'title': 'Book Value/Share'})
        bshval = bsh.next_element.next_element.next_element.next_element
        bshval = bshval.text

        industry = soup.find(attrs={'class':'industry-value'})
        industr =industry.get_text()
    
    
        debtl.append(dval)
        currentl.append(currentval)
        pbl.append(pbval)
        perl.append(perval)
        bshl.append(bshval)
        indl.append(industr)
    except:
        
        #prints the ticker with no data
        print str(n)  + 'here' + a
        pass
        
    
print 'Done'

n=1
for i in b:
    n+=1
    sheet.cell(row=n, column=1).value=i

n=1
for c in debtl:
    n+=1
    sheet.cell(row=n, column=2).value=c
n=1
for d in currentl:
    n+=1
    sheet.cell(row=n, column=3).value=d
n=1
for e in pbl:
    n+=1
    sheet.cell(row=n, column=4).value=e
n=1
for f in perl:
    n+=1
    sheet.cell(row=n, column=5).value=f
n=1
for g in bshl:
    n+=1
    sheet.cell(row=n, column=6).value=g
n=1
for h in indl:
    n+=1
    sheet.cell(row=n, column=7).value=h



wb.save(a+'.xlsx')
wb.close()

print 'Completed'
    
